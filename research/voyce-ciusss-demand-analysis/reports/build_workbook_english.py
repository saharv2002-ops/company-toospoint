"""Build the English-source companion workbook.

Tabs:
    1. Summary
    2. English Rows (raw)
    3. Demand by Pair (all 89 targets, sorted by serviced minutes)
    4. Cancellation Audit (by pair / hour / weekday)
    5. True Demand
    6. Interpreter Sizing (per-pair + shared-pool, three capacity scenarios)
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/home/mocap/toospoint/Voyce data Jan - Apr JEWISH MEMORIAL.xlsx'
OUT = '/home/mocap/toospoint/reports/data/Voyce_English_Analysis.xlsx'
SOURCE = 'english'

NAVY='0B1D3A'; SKY='7FC5E3'; SKY_LITE='E6F4FB'; BONE='FBF8F3'; RULE_GRAY='D9D9D9'
CAP_STD = 6400  # serviced minutes / interpreter / month
CAP_CONS = 4800
CAP_AGG = 8000

df = pd.read_excel(SRC)
df['Request Time (Eastern Standard Time)'] = pd.to_datetime(df['Request Time (Eastern Standard Time)'])
df['hour'] = df['Request Time (Eastern Standard Time)'].dt.hour
df['weekday'] = df['Request Time (Eastern Standard Time)'].dt.day_name()

en_all  = df[df['Source Language'].astype(str).str.strip().str.lower()==SOURCE].copy()
en_serv = en_all[en_all['Status']=='Serviced'].copy()
en_canx = en_all[en_all['Status']=='Cancelled'].copy()

n_total = len(df)
n_en    = len(en_all)
n_serv  = len(en_serv)
n_canx  = len(en_canx)
cx_en   = n_canx/n_en if n_en else 0
cx_fr   = ((df[df['Source Language']=='French']['Status']=='Cancelled').mean())
cx_all  = (df['Status']=='Cancelled').mean()
months  = 4
yr_factor = 12/months

# Demand by pair (serviced)
g = en_serv.groupby('Target Language').agg(
    calls_4mo=('Id','count'),
    serviced_min_4mo=('Service Minutes','sum'),
    avg_min=('Service Minutes','mean'),
    median_min=('Service Minutes','median'),
).reset_index()
g['calls_yr_proj']     = (g['calls_4mo']*yr_factor).round(0).astype(int)
g['serviced_min_yr']   = (g['serviced_min_4mo']*yr_factor).round(0).astype(int)
g['serviced_min_per_mo'] = (g['serviced_min_yr']/12).round(0).astype(int)
g['interpreters_needed_serviced'] = np.ceil(g['serviced_min_per_mo']/CAP_STD).astype(int)
g = g.sort_values('serviced_min_4mo', ascending=False)

# Cancellation audit
cancellation_by_pair = en_all.groupby('Target Language').agg(
    total_calls=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
    serviced=('Status', lambda s:(s=='Serviced').sum()),
).reset_index()
cancellation_by_pair['cancellation_rate'] = (cancellation_by_pair['cancelled']/cancellation_by_pair['total_calls']).round(3)
cancellation_by_pair = cancellation_by_pair.sort_values('total_calls', ascending=False)

cancellation_by_hour = en_all.groupby('hour').agg(
    total=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
).reset_index()
cancellation_by_hour['cancellation_rate']=(cancellation_by_hour['cancelled']/cancellation_by_hour['total']).round(3)

wd_order=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
cancellation_by_wd = en_all.groupby('weekday').agg(
    total=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
).reindex([w for w in wd_order if w in en_all['weekday'].values]).reset_index()
cancellation_by_wd['cancellation_rate']=(cancellation_by_wd['cancelled']/cancellation_by_wd['total']).round(3)

# True demand (cancelled imputed at avg serviced duration per pair)
avg = en_serv.groupby('Target Language')['Service Minutes'].mean()
unmet_count = en_canx.groupby('Target Language')['Id'].count()
unmet_min   = (unmet_count*avg).round(0)

# build a complete frame keyed by all targets that appear in either set
all_targets = sorted(set(en_serv['Target Language'].dropna()) | set(en_canx['Target Language'].dropna()))
true = pd.DataFrame({'Target Language': all_targets})
true['serviced_calls_4mo']    = true['Target Language'].map(en_serv.groupby('Target Language').size()).fillna(0).astype(int)
true['cancelled_calls_4mo']   = true['Target Language'].map(unmet_count).fillna(0).astype(int)
true['avg_serviced_call_min'] = true['Target Language'].map(avg.round(1)).fillna(0)
true['serviced_min_4mo']      = true['Target Language'].map(en_serv.groupby('Target Language')['Service Minutes'].sum()).fillna(0).astype(int)
true['unmet_min_estimate_4mo']= true['Target Language'].map(unmet_min).fillna(0).astype(int)
true['true_demand_min_4mo']   = (true['serviced_min_4mo']+true['unmet_min_estimate_4mo']).astype(int)
true['true_demand_min_yr']    = (true['true_demand_min_4mo']*yr_factor).round(0).astype(int)
true['true_demand_min_per_mo']= (true['true_demand_min_yr']/12).round(0).astype(int)
true['interpreters_needed_TRUE']=np.ceil(true['true_demand_min_per_mo']/CAP_STD).astype(int)
true = true.sort_values('true_demand_min_yr', ascending=False)

# Sizing — three scenarios
sizing = true[['Target Language','true_demand_min_per_mo']].copy()
sizing['conservative_4.8k'] = np.ceil(sizing['true_demand_min_per_mo']/CAP_CONS).astype(int)
sizing['standard_6.4k']     = np.ceil(sizing['true_demand_min_per_mo']/CAP_STD).astype(int)
sizing['aggressive_8.0k']   = np.ceil(sizing['true_demand_min_per_mo']/CAP_AGG).astype(int)

# ─────────────────  WRITE WORKBOOK ─────────────────
wb = Workbook()
ws = wb.active
wb.remove(ws)

THIN = Side(border_style='thin', color=RULE_GRAY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, n_cols, fill=NAVY, font_color='FFFFFF'):
    for c in range(1, n_cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(bold=True, color=font_color, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def style_body(ws, start, n_rows, n_cols, alt=BONE):
    for r in range(start, start+n_rows):
        for c in range(1, n_cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left' if c==1 else 'right', vertical='center')
            cell.border = BORDER
            if (r-start)%2==1:
                cell.fill = PatternFill('solid', fgColor=alt)

def autosize(ws, df, extra=4, max_w=42):
    for i, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 5) + extra
        ws.column_dimensions[get_column_letter(i)].width = min(w, max_w)

def write_df(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=1).font = Font(bold=True, size=14, color=NAVY, name='Calibri')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(len(df.columns), 4))
        r += 2
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=r, column=ci).value = col
    style_header(ws, r, len(df.columns))
    body_start = r+1
    for ri, row in enumerate(df.itertuples(index=False), body_start):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci).value = val
    style_body(ws, body_start, len(df), len(df.columns))
    autosize(ws, df)

# 1) Summary
ws_sum = wb.create_sheet('1. Summary')
ws_sum['A1'] = 'Voyce — English-Source Demand Analysis'
ws_sum['A1'].font = Font(bold=True, size=20, color=NAVY)
ws_sum.merge_cells('A1:D1')
ws_sum['A2'] = 'Client: CIUSSS Centre-Ouest-de-l’Île-de-Montréal (incl. Jewish General Hospital network)'
ws_sum['A2'].font = Font(italic=True, size=11, color='5E6678')
ws_sum.merge_cells('A2:D2')
ws_sum['A3'] = f'Window: Jan 1 - Apr 30, 2025 ({months} months) • Annualization factor x{yr_factor:.1f}'
ws_sum['A3'].font = Font(size=10, color='5E6678')
ws_sum.merge_cells('A3:D3')

kpis = [
    ('Metric','Value'),
    ('Total rows in file', n_total),
    ('English-source calls', n_en),
    ('   - Serviced', n_serv),
    ('   - Cancelled', n_canx),
    ('English-source cancellation rate', f'{cx_en:.1%}'),
    ('French-source cancellation rate (for context)', f'{cx_fr:.1%}'),
    ('Distinct target languages observed', en_all['Target Language'].nunique()),
    ('Serviced English minutes (4 mo)', int(en_serv['Service Minutes'].sum())),
    ('Annualized serviced minutes', int(en_serv['Service Minutes'].sum()*yr_factor)),
    ('TRUE-demand minutes (4 mo, serviced + unmet)', int(true['true_demand_min_4mo'].sum())),
    ('TRUE-demand minutes (annualized)', int(true['true_demand_min_yr'].sum())),
    ('TRUE-demand minutes per month', int(true['true_demand_min_per_mo'].sum())),
    ('Interpreters needed (per-pair, true demand, std cap)', int(true['interpreters_needed_TRUE'].sum())),
    ('Interpreters needed (shared pool, true demand, std cap)', int(np.ceil(true['true_demand_min_per_mo'].sum()/CAP_STD))),
]
start = 5
for ri, row in enumerate(kpis, start):
    for ci, val in enumerate(row, 1):
        ws_sum.cell(row=ri, column=ci).value = val
style_header(ws_sum, start, 2)
style_body(ws_sum, start+1, len(kpis)-1, 2)
ws_sum.column_dimensions['A'].width = 56
ws_sum.column_dimensions['B'].width = 28

# 2) Raw filtered English rows
ws_raw = wb.create_sheet('2. English Rows (raw)')
write_df(ws_raw, en_all.drop(columns=['hour','weekday']), title='All rows where Source Language = English')

# 3) Demand by pair
ws_dem = wb.create_sheet('3. Demand by Pair')
write_df(ws_dem, g, title='Serviced English-source demand by target language (sorted by minutes)')

# 4) Cancellation audit — multi-section sheet
ws_canx = wb.create_sheet('4. Cancellation Audit')
ws_canx['A1']='English-source cancellation analysis'
ws_canx['A1'].font = Font(bold=True, size=14, color=NAVY)
ws_canx.merge_cells('A1:E1')
write_df(ws_canx, cancellation_by_pair, start_row=3, title='By target language')
write_df(ws_canx, cancellation_by_hour, start_row=3+len(cancellation_by_pair)+5, title='By hour of day')
write_df(ws_canx, cancellation_by_wd,   start_row=3+len(cancellation_by_pair)+5+len(cancellation_by_hour)+5, title='By weekday')

# 5) True demand
ws_true = wb.create_sheet('5. True Demand')
write_df(ws_true, true, title='True demand = serviced minutes + cancellations imputed at avg duration')

# 6) Sizing
ws_size = wb.create_sheet('6. Interpreter Sizing')
write_df(ws_size, sizing, title='Per-pair interpreter need under three capacity scenarios (min/mo/interpreter)')

wb.save(OUT)
print('Wrote:', OUT)
