"""Build a multi-tab analysis workbook from the Voyce Jan-Apr file.

Tabs produced:
    1. Raw_Filtered_French   - all rows where Source Language = French
    2. Summary               - high-level KPIs
    3. Demand_by_Pair        - serviced minutes / pair, annualized
    4. Cancellation_Audit    - cancellation rates by pair, hour, weekday
    5. True_Demand_Estimate  - serviced + cancelled-as-unmet
    6. Interpreter_Sizing    - vs current pool, gap analysis
    7. Pool_Reference        - the user's current French pool
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/home/mocap/toospoint/Voyce data Jan - Apr JEWISH MEMORIAL.xlsx'
OUT = '/home/mocap/toospoint/reports/data/Voyce_French_Analysis.xlsx'

NAVY = '0B1D3A'
SKY  = '7FC5E3'
SKY_LITE = 'E6F4FB'
GOLD = 'E2B770'
CORAL = 'EE7D6A'
BONE = 'FBF8F3'
RULE_GRAY = 'D9D9D9'

CURRENT_POOL = {
    'Amharic':2,'Oromo':2,'Somali':1,'Sango':1,'Soninke':1,'Bambara':1,
    'Arabic':7,'Chinese Mandarin':4,'Chinese Cantonese':2,
    'Portuguese':3,'Portuguese Brazilian':3,'Portuguese Continental':2,
    'Wolof':7,'Pulaar':4,'Fulani':3,
    'Kinyarwanda':25,'Swahili':24,'Kirundi':15,'Kinyamulenge':11,
    'Haitian Creole':64,'Spanish':19,'Lingala':22,
    'Albanian':1,'Italian':2,'German':1,
    'Gujarati':1,'Hindi':2,'Punjabi':1,'Urdu':1,
    'Japanese':1,'Tamil':1,'Tamil Sri Lankan':1,
    'Russian':1,'Turkish':1,'Cambodian Khmer':3,
    'Luba Kasai':1,'Yiddish':1,'Hebrew':1,'Yoruba':2,
}

INTERPRETER_CAPACITY_MIN_PER_MO = 6400  # standard

df = pd.read_excel(SRC)
df['Request Time (Eastern Standard Time)'] = pd.to_datetime(df['Request Time (Eastern Standard Time)'])
df['hour'] = df['Request Time (Eastern Standard Time)'].dt.hour
df['weekday'] = df['Request Time (Eastern Standard Time)'].dt.day_name()

# ─── Filtered French ───
fr_all  = df[df['Source Language'].astype(str).str.strip().str.lower()=='french'].copy()
fr_serv = fr_all[fr_all['Status']=='Serviced'].copy()
fr_canx = fr_all[fr_all['Status']=='Cancelled'].copy()

# ─── Summary KPIs ───
n_rows_total      = len(df)
n_french          = len(fr_all)
n_french_serv     = len(fr_serv)
n_french_canx     = len(fr_canx)
cx_rate_all       = (df['Status']=='Cancelled').mean()
cx_rate_french    = n_french_canx / n_french if n_french else 0
cx_rate_english   = ((df[df['Source Language']=='English']['Status']=='Cancelled').mean())
months_in_window  = 4
yr_factor         = 12/months_in_window

# Demand by pair (serviced)
g = fr_serv.groupby('Target Language').agg(
    calls_4mo=('Id','count'),
    serviced_min_4mo=('Service Minutes','sum'),
    avg_min=('Service Minutes','mean'),
    median_min=('Service Minutes','median'),
).reset_index()
g['calls_yr_proj']     = (g['calls_4mo']*yr_factor).round(0).astype(int)
g['serviced_min_yr']   = (g['serviced_min_4mo']*yr_factor).round(0).astype(int)
g['serviced_min_per_mo'] = (g['serviced_min_yr']/12).round(0).astype(int)
g['interpreters_needed_serviced'] = np.ceil(g['serviced_min_per_mo']/INTERPRETER_CAPACITY_MIN_PER_MO).astype(int)

# Cancellation audit
cancellation_by_pair = fr_all.groupby('Target Language').agg(
    total_calls=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
    serviced=('Status', lambda s:(s=='Serviced').sum()),
).reset_index()
cancellation_by_pair['cancellation_rate'] = (cancellation_by_pair['cancelled']/cancellation_by_pair['total_calls']).round(3)

cancellation_by_hour = fr_all.groupby('hour').agg(
    total=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
).reset_index()
cancellation_by_hour['cancellation_rate']=(cancellation_by_hour['cancelled']/cancellation_by_hour['total']).round(3)

wd_order=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
cancellation_by_wd = fr_all.groupby('weekday').agg(
    total=('Id','count'),
    cancelled=('Status', lambda s:(s=='Cancelled').sum()),
).reindex([w for w in wd_order if w in fr_all['weekday'].values]).reset_index()
cancellation_by_wd['cancellation_rate']=(cancellation_by_wd['cancelled']/cancellation_by_wd['total']).round(3)

# True demand (cancellations imputed as unmet at avg duration)
avg_min_per_pair = fr_serv.groupby('Target Language')['Service Minutes'].mean()
unmet_count = fr_canx.groupby('Target Language')['Id'].count()
unmet_min = (unmet_count*avg_min_per_pair).round(0)
true = pd.DataFrame({
    'Target Language': avg_min_per_pair.index,
    'serviced_calls_4mo': fr_serv.groupby('Target Language').size().values,
    'cancelled_calls_4mo': unmet_count.reindex(avg_min_per_pair.index).fillna(0).astype(int).values,
    'avg_serviced_call_min': avg_min_per_pair.round(1).values,
    'serviced_min_4mo': fr_serv.groupby('Target Language')['Service Minutes'].sum().values,
    'unmet_min_estimate_4mo': unmet_min.reindex(avg_min_per_pair.index).fillna(0).astype(int).values,
})
true['true_demand_min_4mo'] = (true['serviced_min_4mo'] + true['unmet_min_estimate_4mo']).astype(int)
true['true_demand_min_yr']  = (true['true_demand_min_4mo']*yr_factor).round(0).astype(int)
true['true_demand_min_per_mo'] = (true['true_demand_min_yr']/12).round(0).astype(int)
true['interpreters_needed_TRUE'] = np.ceil(true['true_demand_min_per_mo']/INTERPRETER_CAPACITY_MIN_PER_MO).astype(int)

# Interpreter sizing — current pool vs needs
sizing = true[['Target Language','true_demand_min_per_mo','interpreters_needed_TRUE']].copy()
sizing['current_pool'] = sizing['Target Language'].map(CURRENT_POOL).fillna(0).astype(int)
sizing['gap_(pool - need)'] = sizing['current_pool'] - sizing['interpreters_needed_TRUE']

# Pool reference table
pool_df = pd.DataFrame(
    sorted(CURRENT_POOL.items(), key=lambda x:-x[1]),
    columns=['Target Language','Current Interpreters']
)

# ───────────────  WRITE WORKBOOK ───────────────
wb = Workbook()
ws = wb.active
wb.remove(ws)

THIN = Side(border_style='thin', color=RULE_GRAY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, n_cols, fill=NAVY, font_color='FFFFFF'):
    for c in range(1, n_cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(bold=True, color=font_color, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER

def style_body(ws, start, n_rows, n_cols, alt=BONE):
    for r in range(start, start+n_rows):
        for c in range(1, n_cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(horizontal='left' if c==1 else 'right', vertical='center')
            cell.border = BORDER
            if (r-start)%2==1:
                cell.fill = PatternFill('solid', fgColor=alt)

def autosize(ws, df, extra=4, max_w=42):
    for i, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 5) + extra
        ws.column_dimensions[get_column_letter(i)].width = min(w, max_w)

def write_df(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=1).font = Font(bold=True, size=14, color=NAVY, name='Calibri')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max(len(df.columns), 4))
        r += 2
    for ci, col in enumerate(df.columns, 1):
        ws.cell(row=r, column=ci).value = col
    style_header(ws, r, len(df.columns))
    body_start = r+1
    for ri, row in enumerate(df.itertuples(index=False), body_start):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci).value = val
    style_body(ws, body_start, len(df), len(df.columns))
    autosize(ws, df)

# 1) Cover/Summary tab
ws_sum = wb.create_sheet('1. Summary')
ws_sum['A1'] = 'Voyce — French-Source Demand Analysis'
ws_sum['A1'].font = Font(bold=True, size=20, color=NAVY)
ws_sum.merge_cells('A1:D1')
ws_sum['A2'] = 'Client: CIUSSS Centre-Ouest-de-l’Île-de-Montréal (incl. Jewish General Hospital network)'
ws_sum['A2'].font = Font(italic=True, size=11, color='5E6678')
ws_sum.merge_cells('A2:D2')
ws_sum['A3'] = f'Window: Jan 1 - Apr 30, 2025 ({months_in_window} months) • Annualization factor x{yr_factor:.1f}'
ws_sum['A3'].font = Font(size=10, color='5E6678')
ws_sum.merge_cells('A3:D3')

kpis = [
    ('Metric','Value'),
    ('Total rows in file', n_rows_total),
    ('French-source calls', n_french),
    ('   - Serviced', n_french_serv),
    ('   - Cancelled', n_french_canx),
    ('French-source cancellation rate', f'{cx_rate_french:.1%}'),
    ('English-source cancellation rate', f'{cx_rate_english:.1%}'),
    ('Overall cancellation rate', f'{cx_rate_all:.1%}'),
    ('Serviced French minutes (4 mo)', int(fr_serv['Service Minutes'].sum())),
    ('Annualized serviced minutes', int(fr_serv['Service Minutes'].sum()*yr_factor)),
    ('TRUE-demand minutes (4 mo, serviced + unmet)', int(true['true_demand_min_4mo'].sum())),
    ('TRUE-demand minutes (annualized)', int(true['true_demand_min_yr'].sum())),
    ('TRUE-demand minutes per month', int(true['true_demand_min_per_mo'].sum())),
    ('Interpreters needed (per-pair, true demand)', int(true['interpreters_needed_TRUE'].sum())),
    ('Interpreters needed (shared pool, true demand)', int(np.ceil(true['true_demand_min_per_mo'].sum()/INTERPRETER_CAPACITY_MIN_PER_MO))),
]
start = 5
for ri, row in enumerate(kpis, start):
    for ci, val in enumerate(row, 1):
        ws_sum.cell(row=ri, column=ci).value = val
style_header(ws_sum, start, 2)
style_body(ws_sum, start+1, len(kpis)-1, 2)
ws_sum.column_dimensions['A'].width = 52
ws_sum.column_dimensions['B'].width = 28

# 2) Raw filtered French rows
ws_raw = wb.create_sheet('2. French Rows (raw)')
write_df(ws_raw, fr_all.drop(columns=['hour','weekday']), title='All rows where Source Language = French')

# 3) Demand by pair
ws_dem = wb.create_sheet('3. Demand by Pair')
write_df(ws_dem, g, title='Serviced French-source demand by target language')

# 4) Cancellation audit
ws_canx = wb.create_sheet('4. Cancellation Audit')
ws_canx['A1']='French-source cancellation analysis'
ws_canx['A1'].font = Font(bold=True, size=14, color=NAVY)
ws_canx.merge_cells('A1:E1')
write_df(ws_canx, cancellation_by_pair, start_row=3, title='By target language')
write_df(ws_canx, cancellation_by_hour, start_row=3+len(cancellation_by_pair)+5, title='By hour of day')
write_df(ws_canx, cancellation_by_wd, start_row=3+len(cancellation_by_pair)+5+len(cancellation_by_hour)+5, title='By weekday')

# 5) True demand
ws_true = wb.create_sheet('5. True Demand')
write_df(ws_true, true, title='True demand = serviced minutes + cancellations imputed at avg duration')

# 6) Interpreter sizing
ws_size = wb.create_sheet('6. Interpreter Sizing')
write_df(ws_size, sizing, title=f'Pool vs need (capacity = {INTERPRETER_CAPACITY_MIN_PER_MO:,} serviced min / interpreter / month)')

# 7) Pool reference
ws_pool = wb.create_sheet('7. Current Pool')
write_df(ws_pool, pool_df, title='Current French interpreter pool (provided)')

wb.save(OUT)
print("Wrote:", OUT)
